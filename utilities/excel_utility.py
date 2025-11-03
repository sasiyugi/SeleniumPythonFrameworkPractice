import openpyxl


def read_data_from_excel(excel_path, sheet_name):

    final_list = []
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    rows = sheet.max_row
    columns = sheet.max_column

    for r in range(2, rows+1):
        row_list = []
        for c in range(1, columns+1):
            row_list.append(sheet.cell(r, c).value)

        final_list.append(row_list)

    return final_list